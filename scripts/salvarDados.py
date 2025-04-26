from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side
import os
from datetime import datetime



def salvar_dados(temperatura, umidade):
    file_path = "dados_temperatura_umidade.xlsx"

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        planilha = wb.active
    else:
        wb = Workbook()
        planilha = wb.active

        planilha["A1"] = "Data"
        planilha["B1"] = "Temperatura"
        planilha["C1"] = "Umidade"

        for celula in planilha[1]:
            celula.font = Font(bold=True, color="FFFFFF")
            celula.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

    agora = datetime.now()
    data_hora = agora.strftime("%d/%m/%Y %H:%M:%S")
    linha_txt = f"{data_hora} - Temperatura: {temperatura} - Umidade: {umidade}\n"
    linha_planilha = [data_hora, temperatura, umidade]

    planilha.append(linha_planilha)

    ultima_linha = planilha[planilha.max_row]
    for celula in ultima_linha:
        celula.font = Font(color="FFFFFF")

    for coluna in planilha.columns:
        max_length = 0
        coluna_letra = get_column_letter(coluna[0].column)

        for celula in coluna:
            try:
                if celula.value:
                    max_length = max(max_length, len(str(celula.value)))
            except:
                pass

        adjusted_width = max_length + 2
        planilha.column_dimensions[coluna_letra].width = adjusted_width

    borda = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    for linha in planilha.iter_rows():
        for celula in linha:
            celula.border = borda

    wb.save(file_path)

    with open("dados_clima.txt", "a", encoding="utf-8") as arquivo:
        arquivo.write(linha_txt)

    print(f"✅ Dados salvos: {data_hora} | Temperatura: {temperatura} | Umidade: {umidade}")