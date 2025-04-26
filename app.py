import os
import tkinter as tk
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side

def capturar_dados_clima():
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")

    driver_path = os.path.join(os.getcwd(), 'drivers', 'msedgedriver.exe')
    driver = webdriver.Edge(service=Service(driver_path), options=options)

    driver.get('https://www.climatempo.com.br')

    try:
        temperatura_element = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.ID, "current-weather-temperature"))
        )
        while temperatura_element.text.strip() == "":
            temperatura_element = driver.find_element(By.ID, "current-weather-temperature")
        temperatura = temperatura_element.text

        umidade_element = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.ID, "current-weather-humidity"))
        )
        while umidade_element.text.strip() == "":
            umidade_element = driver.find_element(By.ID, "current-weather-humidity")
        umidade = umidade_element.text

    except Exception as e:
        print("Erro ao capturar dados:", e)
        temperatura = "N/A"
        umidade = "N/A"
    finally:
        driver.quit()

    return temperatura, umidade

def salvar_dados(temperatura, umidade):
    file_path = "dados_temperatura_umidade.xlsx"

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        planilha = wb.active
    else:
        wb = Workbook()
        planilha = wb.active

        planilha["A1"] = "Data"
        planilha["B1"] = "Temperatura"
        planilha["C1"] = "Umidade"

        for celula in planilha[1]:
            celula.font = Font(bold=True, color="FFFFFF")
            celula.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

    agora = datetime.now()
    data_hora = agora.strftime("%d/%m/%Y %H:%M:%S")
    linha_txt = f"{data_hora} - Temperatura: {temperatura} - Umidade: {umidade}\n"
    linha_planilha = [data_hora, temperatura, umidade]

    planilha.append(linha_planilha)

    ultima_linha = planilha[planilha.max_row]
    for celula in ultima_linha:
        celula.font = Font(color="FFFFFF")

    for coluna in planilha.columns:
        max_length = 0
        coluna_letra = get_column_letter(coluna[0].column)

        for celula in coluna:
            try:
                if celula.value:
                    max_length = max(max_length, len(str(celula.value)))
            except:
                pass

        adjusted_width = max_length + 2
        planilha.column_dimensions[coluna_letra].width = adjusted_width

    borda = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    for linha in planilha.iter_rows():
        for celula in linha:
            celula.border = borda

    wb.save(file_path)

    with open("dados_clima.txt", "a", encoding="utf-8") as arquivo:
        arquivo.write(linha_txt)

    print(f"✅ Dados salvos: {data_hora} | Temperatura: {temperatura} | Umidade: {umidade}")

def executar_automacao():
    global botao_executar 
    botao_executar.config(state=tk.DISABLED) 
    label_status.config(text="Capturando dados...", fg="#FFA500") 

    temperatura, umidade = capturar_dados_clima()
    salvar_dados(temperatura, umidade)

    label_status.config(text=f"Dados capturados e salvos!\nTemperatura: {temperatura} | Umidade: {umidade}", fg="#32CD32")
    botao_executar.config(state=tk.NORMAL)

    print("Processo finalizado com sucesso! Dados capturados e armazenados.")

def configurar_interface():
    global botao_executar 
    janela = tk.Tk()
    janela.title("Captura de Temperatura e Umidade")
    janela.geometry("500x300")  
    janela.config(bg="#E0E0E0")  

    titulo = tk.Label(janela, text="Monitoramento do Clima", font=("Helvetica", 18, "bold"), bg="#E0E0E0", fg="#333")
    titulo.pack(pady=20)

    descricao = tk.Label(janela, text="Clique abaixo para capturar os dados de temperatura e umidade.", 
                         font=("Arial", 12), bg="#E0E0E0", fg="#555")
    descricao.pack(pady=10)

    botao_executar = tk.Button(janela, text="Capturar Dados", font=("Arial", 14), command=executar_automacao, 
                               bg="#1abc9c", fg="white", relief="flat", padx=20, pady=10, bd=0)
    botao_executar.pack(pady=20)

    global label_status
    label_status = tk.Label(janela, text="", font=("Arial", 12), bg="#E0E0E0", fg="#333")
    label_status.pack(pady=15)

    def on_enter(e):
        botao_executar.config(bg="#45a049")

    def on_leave(e):
        botao_executar.config(bg="#4CAF50")

    botao_executar.bind("<Enter>", on_enter)
    botao_executar.bind("<Leave>", on_leave)

    janela.mainloop()

configurar_interface()
